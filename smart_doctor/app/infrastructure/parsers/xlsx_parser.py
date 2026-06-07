"""
XLSX 表格解析器（v2.1）

使用 openpyxl 提取文本，支持：
- 逐工作表解析
- 行数据文本化（| 分隔）
- 多行合并为上下文
"""
from __future__ import annotations

import logging
import os
import time

from app.infrastructure.parsers.base import DocumentParser, ParsedDocument, TextSegment

logger = logging.getLogger(__name__)


class XlsxParser(DocumentParser):
    supported_types = {"xlsx", "xls"}

    async def parse(self, file_path: str, **kwargs) -> ParsedDocument:
        t0 = time.time()
        file_size = _safe_size(file_path)

        if file_path.lower().endswith(".xls") and not file_path.lower().endswith(".xlsx"):
            return _error_doc(
                "旧版 .xls 格式暂不支持，请转为 .xlsx 格式后再上传", file_size, t0
            )

        try:
            import openpyxl
        except ImportError:
            return _error_doc(
                "openpyxl 未安装，请执行 pip install openpyxl", file_size, t0
            )

        segments: list[TextSegment] = []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_count = len(wb.sheetnames)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_rows = kwargs.get("max_rows_per_sheet", 1000)
                row_count = 0

                rows_data: list[str] = []
                for row in ws.iter_rows(values_only=True):
                    if row_count >= max_rows:
                        rows_data.append(f"[工作表 {sheet_name} 行数超过 {max_rows}，已截断]")
                        break

                    cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                    if cells:
                        rows_data.append(" | ".join(cells))
                    row_count += 1

                if rows_data:
                    # 每 20 行合并为一个 segment（上下文密度）
                    chunk_size = 20
                    for ci in range(0, len(rows_data), chunk_size):
                        chunk_lines = rows_data[ci:ci + chunk_size]
                        segments.append(TextSegment(
                            text="\n".join(chunk_lines),
                            sheet=sheet_name,
                            row=ci + 1,
                            heading=f"工作表: {sheet_name}",
                            confidence=0.88,  # 表格数据置信度略低
                        ))

            wb.close()
        except Exception as e:
            return _error_doc(str(e), file_size, t0)

        elapsed = (time.time() - t0) * 1000
        logger.info(
            "XLSX parsed: path=%s sheets=%d segments=%d time=%.0fms",
            file_path, sheet_count, len(segments), elapsed,
        )

        full_text = "\n".join(s.text for s in segments)
        return ParsedDocument(
            text=full_text,
            segments=segments,
            sheet_count=sheet_count,
            encoding="utf-8",
            parse_method="openpyxl",
            parse_duration_ms=elapsed,
            file_size=file_size,
            file_type="xlsx",
        )


def _safe_size(path: str) -> int:
    try:
        return os.path.getsize(path)
    except OSError:
        return 0


def _error_doc(msg: str, file_size: int, t0: float) -> ParsedDocument:
    return ParsedDocument(
        text="",
        parse_method="openpyxl",
        parse_duration_ms=(time.time() - t0) * 1000,
        file_size=file_size,
        file_type="xlsx",
        error=msg,
    )
